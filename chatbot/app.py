import base64
import csv
import json
import os
import time
from io import BytesIO
from threading import Lock

import requests
from dotenv import load_dotenv
from flask import Flask, Response, jsonify, render_template, request, stream_with_context
from werkzeug.utils import secure_filename
from flask.json.provider import DefaultJSONProvider

from knowledge_base import kb

load_dotenv()

# Custom JSON provider for proper UTF-8 encoding
class UTF8JSONProvider(DefaultJSONProvider):
    def dumps(self, obj, **kwargs):
        kwargs.setdefault('ensure_ascii', False)
        kwargs.setdefault('sort_keys', False)
        return super().dumps(obj, **kwargs)

app = Flask(__name__)
app.json = UTF8JSONProvider(app)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB per request
app.config["JSON_AS_ASCII"] = False  # Support non-ASCII characters in JSON
app.config["JSON_SORT_KEYS"] = False

# UTF-8 Response Handler
@app.after_request
def set_charset(response):
    if 'Content-Type' in response.headers:
        if 'application/json' in response.headers['Content-Type']:
            response.headers['Content-Type'] = 'application/json; charset=utf-8'
        elif 'text/html' in response.headers['Content-Type']:
            response.headers['Content-Type'] = 'text/html; charset=utf-8'
    return response

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash").strip()
GEMINI_BASE = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}"
GEMINI_GENERATE_URL = f"{GEMINI_BASE}:generateContent"
GEMINI_STREAM_URL = f"{GEMINI_BASE}:streamGenerateContent"

SYSTEM_PROMPT = (
    "You are a sophisticated, highly capable multilingual AI assistant with expert-level knowledge. "
    "LANGUAGE SUPPORT: You can understand and respond in ALL languages. Always respond in the same language as the user. "
    "Detect the language automatically and maintain that language throughout the conversation.\n"
    "Your responsibilities:\n"
    "1. MULTILINGUAL: Respond fluently in any language the user writes or speaks (English, Hindi, Spanish, French, Chinese, Arabic, etc.)\n"
    "2. CONTEXT AWARENESS: Maintain conversation history and refer back to previous "
    "messages when relevant. Show you understand the full context.\n"
    "3. KNOWLEDGE BASE: When knowledge base context is provided, cite specific sources "
    "and integrate that information naturally into your answers.\n"
    "4. DEEP ANALYSIS: Don't just answer questions—provide comprehensive analysis, "
    "explain your reasoning, and offer multiple perspectives when appropriate.\n"
    "5. STRUCTURED RESPONSES: Use Markdown formatting extensively (headings, lists, "
    "tables, code blocks with language tags) to make complex information clear and scannable.\n"
    "6. PRECISION: Be accurate, specific, and cite sources. When uncertain, say so clearly.\n"
    "7. FILE HANDLING: Analyze uploaded files thoroughly—extract key insights, "
    "summarize content, and relate it to the user's question.\n"
    "Format all answers in clear, professional Markdown. Think step-by-step and "
    "provide thorough, nuanced responses. Always match the user's language."
)

MAX_HISTORY_MESSAGES = 24
MAX_SINGLE_FILE_BYTES = 8 * 1024 * 1024  # 8 MB per file
SESSION_TTL_SECONDS = 6 * 60 * 60  # purge idle sessions after 6 hours
ALLOWED_EXTENSIONS = {"txt", "md", "csv", "pdf", "docx", "xlsx", "xls", "png", "jpg", "jpeg", "webp"}
IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

# session_id -> {"messages": [...], "touched": epoch_seconds}
conversations = {}
conversations_lock = Lock()


def error_response(message, status_code=500):
    return jsonify({"error": message}), status_code


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def mime_type_for(ext: str) -> str:
    return {
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "webp": "image/webp",
    }.get(ext, "application/octet-stream")


def touch_session(session_id):
    """Get (creating if needed) a session's message list and mark it active."""
    now = time.time()
    with conversations_lock:
        # Lazily purge idle sessions so memory doesn't grow forever.
        stale = [sid for sid, data in conversations.items() if now - data["touched"] > SESSION_TTL_SECONDS]
        for sid in stale:
            del conversations[sid]
        if session_id not in conversations:
            conversations[session_id] = {"messages": [], "touched": now}
        conversations[session_id]["touched"] = now
        return conversations[session_id]["messages"]


def append_turn(session_id, user_content, assistant_content):
    with conversations_lock:
        if session_id not in conversations:
            conversations[session_id] = {"messages": [], "touched": time.time()}
        bucket = conversations[session_id]
        bucket["messages"].append({"role": "user", "content": user_content})
        bucket["messages"].append({"role": "assistant", "content": assistant_content})
        bucket["messages"] = bucket["messages"][-MAX_HISTORY_MESSAGES:]
        bucket["touched"] = time.time()


def extract_text_from_file(uploaded_file):
    filename = secure_filename(uploaded_file.filename or "uploaded_file")
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    raw = uploaded_file.read()

    if ext in {"txt", "md"}:
        return raw.decode("utf-8", errors="ignore")[:12000]

    if ext == "csv":
        text = raw.decode("utf-8", errors="ignore")
        rows = list(csv.reader(text.splitlines()))[:60]
        return "\n".join(", ".join(cell for cell in row) for row in rows)[:12000]

    if ext == "pdf":
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(BytesIO(raw))
            pages = []
            for page in reader.pages[:8]:
                pages.append(page.extract_text() or "")
            return "\n".join(pages)[:12000]
        except Exception as exc:
            return f"Could not extract PDF text: {exc}"

    if ext == "docx":
        try:
            from docx import Document
            doc = Document(BytesIO(raw))
            return "\n".join(p.text for p in doc.paragraphs)[:12000]
        except Exception as exc:
            return f"Could not extract DOCX text: {exc}"

    if ext in {"xlsx", "xls"}:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            chunks = []
            for sheet in wb.worksheets[:3]:
                chunks.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(max_row=60, values_only=True):
                    chunks.append(" | ".join("" if v is None else str(v) for v in row))
            return "\n".join(chunks)[:12000]
        except Exception as exc:
            return f"Could not extract Excel text: {exc}"

    return ""


def build_gemini_contents(history, current_parts):
    contents = []
    for msg in history:
        role = "model" if msg.get("role") == "assistant" else "user"
        text = (msg.get("content") or "").strip()
        if text:
            contents.append({"role": role, "parts": [{"text": text}]})
    contents.append({"role": "user", "parts": current_parts})
    return contents


def parse_gemini_reply(result):
    candidates = result.get("candidates", [])
    if not candidates:
        return ""
    parts = candidates[0].get("content", {}).get("parts", [])
    return "".join(part.get("text", "") for part in parts).strip()


def prepare_turn(session_id, user_message, files):
    """Shared setup for both the streaming and non-streaming chat endpoints.

    Returns (payload, visible_user_content, file_summaries, error_message).
    """
    current_parts = []
    visible_user_content = user_message
    
    # Enhance message with knowledge base context
    kb_context = ""
    if user_message:
        kb_context = kb.get_context_for_query(user_message, max_tokens=1500)
    
    if user_message:
        # Add KB context before user message if available
        if kb_context:
            current_parts.append({"text": kb_context})
        current_parts.append({"text": f"User question: {user_message}"})

    file_summaries = []
    for f in files[:5]:
        filename = secure_filename(f.filename or "uploaded_file")
        if not filename or not allowed_file(filename):
            file_summaries.append(f"Skipped unsupported file: {filename}")
            continue

        f.stream.seek(0, os.SEEK_END)
        size = f.stream.tell()
        f.stream.seek(0)
        if size > MAX_SINGLE_FILE_BYTES:
            file_summaries.append(f"Skipped {filename} (over 8 MB limit)")
            continue

        ext = filename.rsplit(".", 1)[-1].lower()
        if ext in IMAGE_EXTENSIONS:
            raw = f.read()
            current_parts.append({
                "inline_data": {
                    "mime_type": mime_type_for(ext),
                    "data": base64.b64encode(raw).decode("utf-8"),
                }
            })
            current_parts.append({"text": f"User uploaded image: {filename}. Analyze it if relevant."})
            file_summaries.append(filename)
        else:
            text = extract_text_from_file(f)
            current_parts.append({"text": f"\n\nUploaded file: {filename}\nExtracted content:\n{text}"})
            file_summaries.append(filename)

    if file_summaries:
        visible_user_content = (visible_user_content + "\n" if visible_user_content else "") + "Attached: " + ", ".join(file_summaries)

    if not current_parts:
        return None, visible_user_content, file_summaries, "Type a message or upload a file."

    history = touch_session(session_id)
    payload = {
        "systemInstruction": {"parts": [{"text": SYSTEM_PROMPT}]},
        "contents": build_gemini_contents(history[-MAX_HISTORY_MESSAGES:], current_parts),
        "generationConfig": {"temperature": 0.7, "maxOutputTokens": 2048},
    }
    return payload, visible_user_content, file_summaries, None


@app.route("/")
def home():
    return render_template("index.html")


@app.route("/chat", methods=["POST"])
def chat():
    """Non-streaming fallback endpoint (used if SSE isn't supported client-side)."""
    if not GEMINI_API_KEY:
        return error_response("Server missing GEMINI_API_KEY. Add it to your .env file.", 500)

    session_id = request.form.get("session_id", "default").strip() or "default"
    user_message = request.form.get("message", "").strip()
    files = request.files.getlist("files")

    payload, visible_user_content, file_summaries, err = prepare_turn(session_id, user_message, files)
    if err:
        return error_response(err, 400)

    try:
        response = requests.post(
            GEMINI_GENERATE_URL,
            params={"key": GEMINI_API_KEY},
            headers={"Content-Type": "application/json"},
            json=payload,
            timeout=60,
        )
        if response.status_code in (400, 401, 403):
            return error_response("Invalid Gemini API key or API access not enabled. Check GEMINI_API_KEY in .env.", response.status_code)
        if response.status_code == 429:
            return error_response("Gemini rate limit reached. Try again later.", 429)
        if response.status_code >= 400:
            return error_response(f"Gemini API error: {response.text}", response.status_code)
        reply = parse_gemini_reply(response.json())
        if not reply:
            return error_response("The API returned an empty reply.", 500)
    except requests.exceptions.Timeout:
        return error_response("API request timed out. Try again.", 504)
    except requests.exceptions.RequestException as exc:
        return error_response(f"API request failed: {exc}", 500)

    append_turn(session_id, visible_user_content, reply)
    return jsonify({"reply": reply, "files": file_summaries})


@app.route("/chat/stream", methods=["POST"])
def chat_stream():
    """Streams the reply back as Server-Sent Events for a real typing effect."""
    if not GEMINI_API_KEY:
        return error_response("Server missing GEMINI_API_KEY. Add it to your .env file.", 500)

    session_id = request.form.get("session_id", "default").strip() or "default"
    user_message = request.form.get("message", "").strip()
    files = request.files.getlist("files")

    payload, visible_user_content, file_summaries, err = prepare_turn(session_id, user_message, files)
    if err:
        return error_response(err, 400)

    def sse(event, data):
        return f"event: {event}\ndata: {json.dumps(data)}\n\n"

    def generate():
        full_reply = []
        try:
            upstream = requests.post(
                GEMINI_STREAM_URL,
                params={"key": GEMINI_API_KEY, "alt": "sse"},
                headers={"Content-Type": "application/json"},
                json=payload,
                stream=True,
                timeout=90,
            )
            if upstream.status_code in (400, 401, 403):
                yield sse("error", {"error": "Invalid Gemini API key or API access not enabled. Check GEMINI_API_KEY in .env."})
                return
            if upstream.status_code == 429:
                yield sse("error", {"error": "Gemini rate limit reached. Try again later."})
                return
            if upstream.status_code >= 400:
                yield sse("error", {"error": f"Gemini API error: {upstream.text[:500]}"})
                return

            for line in upstream.iter_lines(decode_unicode=True):
                if not line or not line.startswith("data:"):
                    continue
                data_str = line[len("data:"):].strip()
                if not data_str or data_str == "[DONE]":
                    continue
                try:
                    chunk = json.loads(data_str)
                except json.JSONDecodeError:
                    continue
                piece = parse_gemini_reply(chunk)
                if piece:
                    full_reply.append(piece)
                    yield sse("delta", {"text": piece})
        except requests.exceptions.Timeout:
            yield sse("error", {"error": "API request timed out. Try again."})
            return
        except requests.exceptions.RequestException as exc:
            yield sse("error", {"error": f"API request failed: {exc}"})
            return

        reply = "".join(full_reply)
        if not reply:
            yield sse("error", {"error": "The API returned an empty reply."})
            return

        append_turn(session_id, visible_user_content, reply)
        yield sse("done", {"reply": reply, "files": file_summaries})

    return Response(stream_with_context(generate()), mimetype="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    })


@app.route("/reset", methods=["POST"])
def reset():
    data = request.get_json(silent=True) or {}
    session_id = (data.get("session_id") or "default").strip()
    with conversations_lock:
        conversations.pop(session_id, None)
    return jsonify({"status": "cleared"})


# ============= KNOWLEDGE BASE ENDPOINTS =============

@app.route("/kb/add", methods=["POST"])
def kb_add():
    """Add a document to the knowledge base."""
    title = request.form.get("title", "").strip()
    content = request.form.get("content", "").strip()
    category = request.form.get("category", "general").strip()
    
    if not title or not content:
        return error_response("Title and content are required", 400)
    
    result = kb.add_document(title, content, category)
    status_code = 200 if result.get("success") else 400
    return jsonify(result), status_code


@app.route("/kb/search", methods=["POST"])
def kb_search():
    """Search the knowledge base."""
    query = request.form.get("query", "").strip()
    
    if not query:
        return error_response("Query is required", 400)
    
    results = kb.search(query, limit=10)
    return jsonify({"query": query, "results": results, "count": len(results)})


@app.route("/kb/list", methods=["GET"])
def kb_list():
    """List all documents in the knowledge base."""
    documents = kb.get_all_documents()
    return jsonify({"documents": documents, "count": len(documents)})


@app.route("/kb/delete/<doc_id>", methods=["POST"])
def kb_delete(doc_id):
    """Delete a document from the knowledge base."""
    result = kb.delete_document(doc_id)
    status_code = 200 if result.get("success") else 404
    return jsonify(result), status_code


@app.route("/kb/stats", methods=["GET"])
def kb_stats():
    """Get knowledge base statistics."""
    stats = kb.get_stats()
    return jsonify(stats)


if __name__ == "__main__":
    app.run(debug=True)
